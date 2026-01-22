#!/usr/bin/env python3
"""
Generate Reasoning for Label-Only Dataset
Uses GPT-4o via OpenRouter or simple rules to generate reasoning for 12K samples
"""

import json
import argparse
import os
from pathlib import Path
from tqdm import tqdm
import openpyxl
import requests
import time

def load_xlsx_dataset(xlsx_path):
    """Load dataset from Excel file."""
    print(f"Loading dataset from: {xlsx_path}")
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Load data
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        data.append(item)
    
    print(f"Loaded {len(data)} samples")
    print(f"Columns: {headers}")
    
    return data

def generate_reasoning_with_gpt4(content, title, label, api_key):
    """Generate reasoning using GPT-4o via OpenRouter."""
    
    prompt = f"""Analyze this Indonesian news article and provide a brief reasoning (max 200 characters) for why it is classified as "{label}".

Title: {title}
Content: {content[:500]}...

Classification: {label}

Provide reasoning in Bahasa Indonesia that explains:
- Tone of the article (positive/neutral/negative)
- Whether it promotes a product/brand
- Whether it uses persuasive language
- Whether it presents multiple perspectives or just one viewpoint

Keep it under 200 characters."""

    try:
        response = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            },
            json={
                "model": "openai/gpt-4o",
                "messages": [
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 100,
                "temperature": 0.3
            },
            timeout=30
        )
        
        if response.status_code == 200:
            reasoning = response.json()['choices'][0]['message']['content'].strip()
            # Truncate if too long
            # if len(reasoning) > 150:
                # reasoning = reasoning[:147] + "..."
            return reasoning
        else:
            # Fallback to simple reasoning
            return generate_simple_reasoning(label)
    except Exception as e:
        print(f"Warning: GPT-4o failed ({e}), using fallback")
        return generate_simple_reasoning(label)

def generate_simple_reasoning(label):
    """Generate simple reasoning based on label (fallback)."""
    if label == "native ads":
        return "Artikel memiliki nada positif dan mempromosikan produk/brand tertentu dengan bahasa persuasif."
    else:
        return "Artikel bersifat netral dan objektif, menyajikan informasi tanpa promosi produk/brand."

def generate_dataset_with_reasoning(input_path, output_path, use_gpt4=False, api_key=None):
    """Generate reasoning for all samples."""
    
    # Load original dataset
    data = load_xlsx_dataset(input_path)
    
    # Check API key if using GPT-4
    if use_gpt4 and not api_key:
        api_key = os.getenv('OPENROUTER_API_KEY')
        if not api_key:
            print("⚠️  No OPENROUTER_API_KEY found, using simple reasoning instead")
            use_gpt4 = False
    
    # Generate reasoning for each sample
    print(f"\nGenerating reasoning using: {'GPT-4o' if use_gpt4 else 'Simple rules'}...")
    if use_gpt4:
        print(f"⏱️  Estimated time: ~{len(data) * 2 / 60:.1f} minutes")
        print(f"💰 Estimated cost: ~${len(data) * 0.01:.2f}")
    
    formatted_data = []
    
    for i, item in enumerate(tqdm(data)):
        # Extract fields (adjust column names as needed)
        title = item.get('title', '') or item.get('Title', '') or item.get('judul', '')
        content = item.get('content', '') or item.get('Content', '') or item.get('konten', '')
        label = item.get('label', '') or item.get('Label', '') or item.get('klasifikasi', '')
        
        # Generate reasoning
        if use_gpt4:
            reasoning = generate_reasoning_with_gpt4(content, title, label, api_key)
            time.sleep(0.5)  # Rate limiting
        else:
            reasoning = generate_simple_reasoning(label)
        
        # Format for training
        output = {
            "label": label,
            "confidence": 0.85,
            "reasoning": reasoning
        }
        
        formatted_item = {
            "id": i,
            "instruction": "Klasifikasikan artikel berikut sebagai native ads atau berita murni.",
            "title": title,
            "input": content,
            "output": json.dumps(output, ensure_ascii=False)
        }
        
        formatted_data.append(formatted_item)
    
    # Check label distribution
    from collections import Counter
    labels = Counter()
    for item in formatted_data:
        output = json.loads(item['output'])
        labels[output['label']] += 1
    
    print(f"\nLabel distribution:")
    for label, count in labels.most_common():
        print(f"  {label}: {count} ({count/len(formatted_data)*100:.1f}%)")
    
    # Save
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(formatted_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Dataset saved to: {output_path}")
    print(f"   Total samples: {len(formatted_data)}")
    
    return formatted_data

def main():
    parser = argparse.ArgumentParser(description='Generate reasoning for label-only dataset')
    parser.add_argument('--input', type=str,
                       default='../native_ads_dataset.xlsx',
                       help='Input Excel file')
    parser.add_argument('--output', type=str,
                       default='../data/llm_dataset_12k_with_reasoning.json',
                       help='Output JSON file')
    parser.add_argument('--use-gpt4', action='store_true',
                       help='Use GPT-4o via OpenRouter (requires API key)')
    parser.add_argument('--api-key', type=str,
                       help='OpenRouter API key (or set OPENROUTER_API_KEY env var)')
    
    args = parser.parse_args()
    
    generate_dataset_with_reasoning(
        args.input, 
        args.output, 
        use_gpt4=args.use_gpt4,
        api_key=args.api_key
    )
    
    print(f"\n🚀 Next step:")
    print(f"   python3 finetune.py --dataset {args.output} --max-steps 6000")

if __name__ == "__main__":
    main()
